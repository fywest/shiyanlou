# -*- coding utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
import datetime




list_combine=[]

list_2013 = []
list_2014 = []
list_2015 = []
list_2016 = []



def combine():
    wb = load_workbook('courses.xlsx')
    ws_student = wb['students']

    dict_studentAmount = {}
    for row in ws_student:
        dict_studentAmount[row[1].value] = row[2].value

    ws_time = wb['time']
    ws_combine = wb.create_sheet('combine', 2)

    for row in ws_time:
        dict_combine = {}
        dict_combine['start_time'] = row[0].value
        dict_combine['title'] = row[1].value
        dict_combine['total_time'] = row[2].value
        dict_combine['quantity'] = dict_studentAmount[row[1].value]
        list_combine.append(dict_combine)

    # print(len(list_combine))
    for index in range(1,len(list_combine) + 1):
        dict_tmp = list_combine[index - 1]
        ws_combine.cell(column=1, row=index, value=dict_tmp['start_time'])
        ws_combine.cell(column=2, row=index, value=dict_tmp['title'])
        ws_combine.cell(column=3, row=index, value=dict_tmp['quantity'])
        ws_combine.cell(column=4, row=index, value=dict_tmp['total_time'])
    # print(ws_combine.max_row)
    wb.save('courses.xlsx')



def split():

    list_2013.append(list_combine[0])
    list_2014.append(list_combine[0])
    list_2015.append(list_combine[0])
    list_2016.append(list_combine[0])

    for index in range(2, len(list_combine) + 1):
        dict_tmp={}
        dict_tmp = list_combine[index - 1]
        #print(dict_tmp['start_time'])
        str_year=str(dict_tmp['start_time'])[:4]
        if str_year=='2013':
            list_2013.append(dict_tmp)
        elif str_year=='2014':
            list_2014.append(dict_tmp)
        elif str_year=='2015':
            list_2015.append(dict_tmp)
        elif str_year=='2016':
            list_2016.append(dict_tmp)

    years=['2013','2014','2015','2016']
    for year in years:
        file_save(year)

def file_save(year):
    if year=='2013':
        list_name=list_2013
    elif year=='2014':
        list_name = list_2014
    elif year=='2015':
        list_name = list_2015
    elif year=='2016':
        list_name = list_2016

    wb_tmp=Workbook()
    ws_tmp=wb_tmp.active
    ws_tmp.title=year

    for index in range(1,len(list_name) + 1):
        dict_tmp = list_name[index - 1]
        ws_tmp.cell(column=1, row=index, value=dict_tmp['start_time'])
        ws_tmp.cell(column=2, row=index, value=dict_tmp['title'])
        ws_tmp.cell(column=3, row=index, value=dict_tmp['quantity'])
        ws_tmp.cell(column=4, row=index, value=dict_tmp['total_time'])
    # print(ws_tmp.max_row)
    wb_tmp.save(filename=year+'.xlsx')


if __name__=='__main__':
    combine()
    split()
